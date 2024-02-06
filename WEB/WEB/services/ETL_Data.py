from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import json


def Genera_JSON_Excel_Valido( Path_Excel , fila_inicial , Path_Excel_Salida , Path_JSON_Salida , Lista_Archivos_Disponibles ):
	
	try:

		wb = load_workbook( Path_Excel )
		hoja_activa = wb.active
		
		relleno_verde = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
		relleno_rojo = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
		relleno_blanco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')	

		# Rellenamos de color blanco las casillas ------------------------->>>>>>>>>>>>>>>
		num_fila = fila_inicial
		for fila in hoja_activa.iter_rows( min_row=num_fila , values_only=True ):
			for letra in ["A","B","C","D"]:
				hoja_activa[f'{letra}{num_fila}'].fill = relleno_blanco
			num_fila += 1
		# ----------------------------------------------------------------->>>>>>>>>>>>>>>
		
		num_fila = fila_inicial
		Lista_Codigos_Validos = []
		for fila in hoja_activa.iter_rows( min_row=num_fila , values_only=True ):
			try:
				CODIGO_MENSAJE = hoja_activa[f'A{num_fila}'] #REQUERIDO
				if CODIGO_MENSAJE.value != None: #Paramos ejecusion, fila no valida
					Lista_Codigos_Validos.append( CODIGO_MENSAJE.value )
			except:
				pass
			num_fila += 1

		num_fila = fila_inicial
		Dicc_Mensajes_Respuesta_Automatica = {}
		for fila in hoja_activa.iter_rows( min_row=num_fila , values_only=True ):
			
			Data_Valida = True			
			
			try:

				dicc_fila = {}
				
				CODIGO_MENSAJE = hoja_activa[f'A{num_fila}'] #REQUERIDO 
				
				if CODIGO_MENSAJE.value == None: #Paramos ejecusion, fila no valida
					CODIGO_MENSAJE.fill = relleno_rojo
					Data_Valida = False
				else:
					
					CODIGO_MENSAJE.fill = relleno_verde
					dicc_fila["codigo_mensaje"] = CODIGO_MENSAJE.value


					MENSAJE_RESPUESTA = hoja_activa[f'B{num_fila}'] #Password de Usuario
					MENSAJE_RESPUESTA.fill = relleno_verde #Ponemos en Verde Casilla Password
					dicc_fila["mensaje_respuesta"] = ""
					if MENSAJE_RESPUESTA.value != None:
						dicc_fila["mensaje_respuesta"] = str( MENSAJE_RESPUESTA.value )


					ARCHIVO_MENSAJE = hoja_activa[f'C{num_fila}'] #REQUERIDO @@@@>>>>
					try:
						if ARCHIVO_MENSAJE.value == None:
							pass
						elif ARCHIVO_MENSAJE.value in Lista_Archivos_Disponibles:
							dicc_fila[ "archivo_mensaje" ] = ARCHIVO_MENSAJE.value
						else:
							raise
						ARCHIVO_MENSAJE.fill = relleno_verde
					except:
						Data_Valida = False
						ARCHIVO_MENSAJE.fill = relleno_rojo

					REDIRECCIONAR_A_CODIGO = hoja_activa[f'D{num_fila}'] #REQUERIDO @@@@>>>>
					try:
						if REDIRECCIONAR_A_CODIGO.value == None:
							pass
						elif REDIRECCIONAR_A_CODIGO.value in Lista_Codigos_Validos:
							dicc_fila[ "redireccionar_a_codigo" ] = REDIRECCIONAR_A_CODIGO.value
						else:
							raise
						REDIRECCIONAR_A_CODIGO.fill = relleno_verde
					except:
						Data_Valida = False
						REDIRECCIONAR_A_CODIGO.fill = relleno_rojo

			except:
				Data_Valida = False
			
			print( f"{num_fila} - {Data_Valida}")

			if Data_Valida == True:
				
				Data_Mensje = { "mensaje_respuesta" : dicc_fila[ "mensaje_respuesta" ] }
				
				if "archivo_mensaje" in dicc_fila.keys():
					Data_Mensje[ "archivo_mensaje" ] = dicc_fila[ "archivo_mensaje" ]
				if "redireccionar_a_codigo" in dicc_fila.keys():
					Data_Mensje[ "redireccionar_a_codigo" ] = dicc_fila[ "redireccionar_a_codigo" ]

				Dicc_Mensajes_Respuesta_Automatica[ dicc_fila["codigo_mensaje"] ] = Data_Mensje

			num_fila += 1

	except Exception as error:
		Dicc_Mensajes_Respuesta_Automatica = {}

	wb.save( Path_Excel_Salida )  # Guarda el archivo con un nuevo nombre o la misma ruta para sobrescribirlo
	
	with open( Path_JSON_Salida , "w" ) as fileJSON: # Guarda el JSON con Data
		json.dump( Dicc_Mensajes_Respuesta_Automatica , fileJSON )


def Genera_JSON_Excel_CronoJobs_Valido( Path_Excel , fila_inicial , Path_Excel_Salida , Path_JSON_Salida , Lista_Archivos_Disponibles ):
	
	try:

		wb = load_workbook( Path_Excel )
		hoja_activa = wb.active
		
		relleno_verde = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
		relleno_rojo = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
		relleno_blanco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')	

		# Rellenamos de color blanco las casillas ------------------------->>>>>>>>>>>>>>>
		num_fila = fila_inicial
		for fila in hoja_activa.iter_rows( min_row=num_fila , values_only=True ):
			for letra in ["A","B","C"]:
				hoja_activa[f'{letra}{num_fila}'].fill = relleno_blanco
			num_fila += 1
		# ----------------------------------------------------------------->>>>>>>>>>>>>>>

		num_fila = fila_inicial
		Dicc_Mensajes_Respuesta_Automatica = {}
		for fila in hoja_activa.iter_rows( min_row=num_fila , values_only=True ):
			
			Data_Valida = True
			
			try:

				dicc_fila = {}
				
				ID_USUARIO = hoja_activa[f'A{num_fila}'] #REQUERIDO 
				
				if ID_USUARIO.value == None: #Paramos ejecusion, fila no valida
					ID_USUARIO.fill = relleno_rojo
					Data_Valida = False
				else:
					
					ID_USUARIO.fill = relleno_verde
					dicc_fila["id_usuario"] = ID_USUARIO.value


					MENSAJE = hoja_activa[f'B{num_fila}'] #Password de Usuario
					MENSAJE.fill = relleno_verde #Ponemos en Verde Casilla Password
					dicc_fila["mensaje"] = ""
					if MENSAJE.value != None:
						dicc_fila["mensaje"] = str( MENSAJE.value )


					ARCHIVO_MENSAJE = hoja_activa[f'C{num_fila}'] #REQUERIDO @@@@>>>>
					try:
						if ARCHIVO_MENSAJE.value == None:
							pass
						elif ARCHIVO_MENSAJE.value in Lista_Archivos_Disponibles:
							dicc_fila[ "archivo_mensaje" ] = ARCHIVO_MENSAJE.value
						else:
							raise
						ARCHIVO_MENSAJE.fill = relleno_verde
					except:
						Data_Valida = False
						ARCHIVO_MENSAJE.fill = relleno_rojo


			except:
				Data_Valida = False

			if Data_Valida == True:
				
				Data_Mensje = { "mensaje" : dicc_fila[ "mensaje" ] }
				
				if "archivo_mensaje" in dicc_fila.keys():
					Data_Mensje[ "archivo_mensaje" ] = dicc_fila[ "archivo_mensaje" ]

				Dicc_Mensajes_Respuesta_Automatica[ dicc_fila["id_usuario"] ] = Data_Mensje

			num_fila += 1

	except Exception as error:
		Dicc_Mensajes_Respuesta_Automatica = {}

	wb.save( Path_Excel_Salida )  # Guarda el archivo con un nuevo nombre o la misma ruta para sobrescribirlo
	
	with open( Path_JSON_Salida , "w" ) as fileJSON: # Guarda el JSON con Data
		json.dump( Dicc_Mensajes_Respuesta_Automatica , fileJSON )


if __name__ == "__main__":
	'''
	# Config -------->>>>>>>>
	Path_Excel = 'Respuestas_Automaticas.xlsx'
	fila_inicial = 2
	# --------------->>>>>>>>
	
	Genera_JSON_Excel_Valido( Path_Excel , fila_inicial , 'Respuestas_Automaticas_Estado.xlsx' , "Respuestas_Automaticas.json" , [] )
	'''

	# Config -------->>>>>>>>
	Path_Excel = "C:/Users/54261/Desktop/Telegram_Bots/Cronojobs.xlsx"
	fila_inicial = 2
	Path_Excel_Salida = "C:/Users/54261/Desktop/Telegram_Bots/Salida.xlsx"
	Path_JSON_Salida = "C:/Users/54261/Desktop/Telegram_Bots/Salida.json"
	# --------------->>>>>>>>

	Genera_JSON_Excel_CronoJobs_Valido( Path_Excel , fila_inicial , Path_Excel_Salida , Path_JSON_Salida , ['File 1','logoFarmaceutica.jpg'] )